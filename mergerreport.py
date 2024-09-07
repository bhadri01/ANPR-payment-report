import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime

def find_column(data, possible_names):
    """Utility function to find the closest matching column from possible names."""
    for name in possible_names:
        if name in data.columns:
            return name
    return None

def process_and_generate_excel(input_file, output_file, generate_daily=False, generate_monthly=False, start_date=None, end_date=None):
    # Load the CSV file (without skipping rows)
    try:
        # Read the CSV file with the first row as a header
        data = pd.read_csv(input_file, low_memory=False)
    except Exception as e:
        print(f"Error reading {input_file}: {e}")
        return

    # Find the correct column for 'Challan Date'
    challan_date_column = find_column(data, ['Challan Date'])
    if not challan_date_column:
        print(f"Challan Date column not found in the CSV. Exiting.")
        return

    # Ensure the 'Challan Date' column is properly parsed as a date
    data[challan_date_column] = pd.to_datetime(data[challan_date_column], errors='coerce')

    # Filter data if custom date range is provided
    if start_date and end_date:
        data = data[(data[challan_date_column].dt.date >= start_date) & (data[challan_date_column].dt.date <= end_date)]

    # Initialize the report DataFrame
    report_data = pd.DataFrame()

    # Aggregating the data by date
    report_data['Total Number of Cases'] = data.groupby(data[challan_date_column].dt.date).size()

    # Calculate the number of completed cases (not pending)
    report_data['Number of Cases Completed'] = data[data['Challan Status'] != 'Pending'].groupby(data[challan_date_column].dt.date).size()

    # Calculate the number of pending cases
    report_data['Total Cases Pending'] = data[data['Challan Status'] == 'Pending'].groupby(data[challan_date_column].dt.date).size()

    # Add details for cases in 100, 200-900, and 1000 categories
    report_data['Total No. of Cases in 100'] = data[data['Challan Amount'] == 100].groupby(data[challan_date_column].dt.date).size()
    report_data['Total No. of Cases in 200-900'] = data[(data['Challan Amount'] >= 200) & (data['Challan Amount'] <= 900)].groupby(data[challan_date_column].dt.date).size()
    report_data['Total No. of Cases in 1000'] = data[data['Challan Amount'] == 1000].groupby(data[challan_date_column].dt.date).size()

    # Count the number of cases where fine has been collected for each category
    report_data['Total No. of 100\'s Collected'] = data[(data['Challan Amount'] == 100) & (data['Challan Status'] != 'Pending')].groupby(data[challan_date_column].dt.date).size()
    report_data['Total No. of 200-900\'s Collected'] = data[(data['Challan Amount'] >= 200) & (data['Challan Status'] != 'Pending') & (data['Challan Amount'] <= 900)].groupby(data[challan_date_column].dt.date).size()
    report_data['Total No. of 1000\'s Collected'] = data[(data['Challan Amount'] == 1000) & (data['Challan Status'] != 'Pending')].groupby(data[challan_date_column].dt.date).size()

    # Calculate the total number of cases with fines collected
    report_data['Total Fine Collected (No. of Cases)'] = (
        report_data['Total No. of 100\'s Collected'] +
        report_data['Total No. of 200-900\'s Collected'] +
        report_data['Total No. of 1000\'s Collected']
    )

    # Calculate the collected fine amounts for each fine category
    report_data['Collected Fine Amount in 100'] = data[(data['Challan Amount'] == 100) & (data['Challan Status'] != 'Pending')].groupby(data[challan_date_column].dt.date)['Challan Amount'].sum()
    report_data['Collected Fine Amount in 200-900'] = data[(data['Challan Amount'] >= 200) & (data['Challan Amount'] <= 900) & (data['Challan Status'] != 'Pending')].groupby(data[challan_date_column].dt.date)['Challan Amount'].sum()
    report_data['Collected Fine Amount in 1000'] = data[(data['Challan Amount'] == 1000) & (data['Challan Status'] != 'Pending')].groupby(data[challan_date_column].dt.date)['Challan Amount'].sum()

    # Calculate total amount collected by summing up the amounts for all fine categories
    report_data['Total Amount Collected'] = (
        report_data['Collected Fine Amount in 100'] +
        report_data['Collected Fine Amount in 200-900'] +
        report_data['Collected Fine Amount in 1000']
    )

    # Replace NaN with 0 in case of missing values
    report_data.fillna(0, inplace=True)

    # After processing, aggregate the data by date
    if generate_monthly:
        report_data.index = pd.to_datetime(report_data.index)
        report_data = report_data.groupby([report_data.index.to_period('M')]).sum()

    final_report = report_data.groupby(report_data.index).sum()

    # Convert PeriodIndex to string format for Excel compatibility
    if isinstance(final_report.index, pd.PeriodIndex):
        final_report.index = final_report.index.astype(str)

    # Prepare the final Excel report in the desired format
    wb = Workbook()
    ws = wb.active
    ws.title = 'Report'

    # Merge cells for section headers
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=5)
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=9)
    ws.merge_cells(start_row=1, start_column=10, end_row=1, end_column=13)
    ws.merge_cells(start_row=1, start_column=14, end_row=1, end_column=17)

    # Set headings for each section
    ws['A2'] = 'Date'
    ws['B1'] = 'Total Number of Cases'
    ws['B2'] = 'No. of 100\'s'
    ws['C2'] = 'No. of 200-900\'s'
    ws['D2'] = 'No. of 1000\'s'
    ws['E2'] = 'Total Cases'

    ws['F1'] = 'Total Cases Fine Collected'
    ws['F2'] = 'No. of 100\'s Collected'
    ws['G2'] = 'No. of 200-900\'s Collected'
    ws['H2'] = 'No. of 1000\'s Collected'
    ws['I2'] = 'Total Fine Collected (No. of Cases)'

    ws['J1'] = 'Total Cases Pending'
    ws['J2'] = 'No. of 100\'s Pending'
    ws['K2'] = 'No. of 200-900\'s Pending'
    ws['L2'] = 'No. of 1000\'s Pending'
    ws['M2'] = 'Total Cases Pending (No. of Cases)'

    ws['N1'] = 'Total Amount Collected'
    ws['N2'] = '100\'s Collected'
    ws['O2'] = '200-900\'s Collected'
    ws['P2'] = '1000\'s Collected'
    ws['Q2'] = 'Grand Total (Amount Collected)'

    # Align the merged cells in the center
    for cell in ['B1', 'F1', 'J1', 'N1']:
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    # Prepare rows of data by column names using `iterrows()`
    for index, row in final_report.iterrows():
        grand_total = row['Collected Fine Amount in 100'] + row['Collected Fine Amount in 200-900'] + row['Collected Fine Amount in 1000']
        total_number_of_cases = row['Total No. of Cases in 100'] + row['Total No. of Cases in 200-900'] + row['Total No. of Cases in 1000']
        total_fine_collected = row['Total No. of 100\'s Collected'] + row['Total No. of 200-900\'s Collected'] + row['Total No. of 1000\'s Collected']
        total_pending = (row['Total No. of Cases in 100'] - row['Total No. of 100\'s Collected']) + (row['Total No. of Cases in 200-900'] - row['Total No. of 200-900\'s Collected']) + (row['Total No. of Cases in 1000'] - row['Total No. of 1000\'s Collected'])
        ws.append([
            index,  # Date or Month (depending on report type)
            row['Total No. of Cases in 100'], row['Total No. of Cases in 200-900'], row['Total No. of Cases in 1000'], total_number_of_cases,  # Total Number of Cases
            row['Total No. of 100\'s Collected'], row['Total No. of 200-900\'s Collected'], row['Total No. of 1000\'s Collected'], total_fine_collected,  # Total Cases Fine Collected (No. of Cases)
            row['Total No. of Cases in 100'] - row['Total No. of 100\'s Collected'],  # Pending in 100s
            row['Total No. of Cases in 200-900'] - row['Total No. of 200-900\'s Collected'],  # Pending in 200-900s
            row['Total No. of Cases in 1000'] - row['Total No. of 1000\'s Collected'],  # Pending in 1000s
            total_pending,  # Total Cases Pending (No. of Cases)
            row['Collected Fine Amount in 100'], row['Collected Fine Amount in 200-900'], row['Collected Fine Amount in 1000'], grand_total  # Grand Total (Amount Collected)
        ])

    # Calculate the total for each column and append it to the Excel sheet as the last row
    total_fine_collected_sum = (
        final_report['Total No. of 100\'s Collected'].sum() +
        final_report['Total No. of 200-900\'s Collected'].sum() +
        final_report['Total No. of 1000\'s Collected'].sum()
    )

    grand_total_sum = (
        final_report['Collected Fine Amount in 100'].sum() +
        final_report['Collected Fine Amount in 200-900'].sum() +
        final_report['Collected Fine Amount in 1000'].sum()
    )

    total_row = [
        'Total',  # Label for the totals row
        final_report['Total No. of Cases in 100'].sum(),  # Total for 100's cases
        final_report['Total No. of Cases in 200-900'].sum(),  # Total for 200-900's cases
        final_report['Total No. of Cases in 1000'].sum(),  # Total for 1000's cases
        final_report['Total Number of Cases'].sum(),  # Total Number of Cases

        final_report['Total No. of 100\'s Collected'].sum(),  # Total 100's Collected
        final_report['Total No. of 200-900\'s Collected'].sum(),  # Total 200-900's Collected
        final_report['Total No. of 1000\'s Collected'].sum(),  # Total 1000's Collected
        total_fine_collected_sum,  # Total Fine Collected (No. of Cases)

        final_report['Total No. of Cases in 100'].sum() - final_report['Total No. of 100\'s Collected'].sum(),  # Pending in 100s
        final_report['Total No. of Cases in 200-900'].sum() - final_report['Total No. of 200-900\'s Collected'].sum(),  # Pending in 200-900s
        final_report['Total No. of Cases in 1000'].sum() - final_report['Total No. of 1000\'s Collected'].sum(),  # Pending in 1000s
        final_report['Total Cases Pending'].sum(),  # Total Pending Cases

        final_report['Collected Fine Amount in 100'].sum(),  # Total Fine Collected in 100's
        final_report['Collected Fine Amount in 200-900'].sum(),  # Total Fine Collected in 200-900's
        final_report['Collected Fine Amount in 1000'].sum(),  # Total Fine Collected in 1000's
        grand_total_sum  # Grand Total Amount Collected
    ]

    # Append the totals row to the Excel sheet
    ws.append(total_row)

    # Save the final report to an Excel file
    wb.save(output_file)
    print(f"Final Excel report saved at: {output_file}")

def main():
    print("=== ANPR Fine Details Processing Tool ===")
    print("Please choose an option:")
    print("1. Generate Daily Details Report")
    print("2. Generate Monthly Summary Report")
    print("3. Generate Report for Custom Date Range")
    print("4. Exit")

    choice = input("Enter your choice (1-4): ")

    if choice == '4':
        print("Exiting the tool.")
        return

    input_file = input("Enter the path to the CSV file: ")

    generate_daily = False
    generate_monthly = False
    start_date = None
    end_date = None

    if choice == '1':
        generate_daily = True
    elif choice == '2':
        generate_monthly = True
    elif choice == '3':
        start_date_str = input("Enter start date for custom report (YYYY-MM-DD): ")
        end_date_str = input("Enter end date for custom report (YYYY-MM-DD): ")
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            print("Invalid date format. Please use YYYY-MM-DD.")
            return
    else:
        print("Invalid choice. Please select a valid option.")
        return

    # Set the output file name based on the user's selection
    output_file = 'final_report.xlsx'

    # Call the processing function with the specified options
    process_and_generate_excel(input_file, output_file, generate_daily, generate_monthly, start_date, end_date)

if __name__ == "__main__":
    main()
